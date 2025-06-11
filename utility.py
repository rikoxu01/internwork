# -*- coding:utf-8 -*-

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import optimize


# Daily Yield Curve and Interpolation
def HermiteInter(t, y, xx):
    n = len(t)
    M = []
    for seg in range(1, n+1):
        if seg == 1:
            m = ((t[1] + t[2] - 2 * t[0]) * (y[1] - y[0])/(t[1] - t[0]) - (t[1] - t[0]) * (y[2] - y[1])/(t[2] - t[1])) \
                 / (t[2] - t[0])
        elif seg < n:
            m = ((y[seg - 1] - y[seg - 2]) * (t[seg] - t[seg - 1])/(t[seg - 1] - t[seg-2]) + (y[seg] - y[seg - 1]) *
                 (t[seg - 1] - t[seg-2])/(t[seg] - t[seg - 1]))/(t[seg] - t[seg-2])
        else:
            # m = ((2 * t[n-1] - t[n-2] - t[n-3]) * (y[n-1] - y[n-2]) / (t[n-1] - t[n-2]) - ((y[n-2] - y[n-3]) *
                 # (t[n-1] - t[n-2])) / (t[n-2] - t[n-3])) / (t[n-1] - t[n-3])
            m = 0
        M.append(m)

    yy = np.zeros(shape=np.shape(xx))
    for seg in range(1, n):
        a = y[seg-1]
        b = M[seg-1]
        c = 3 * (y[seg] - y[seg-1])/pow((t[seg] - t[seg-1]), 2) - (M[seg-1] * 2 + M[seg])/(t[seg] - t[seg-1])
        d = (M[seg] + M[seg-1])/pow(t[seg] - t[seg-1], 2) - 2 * (y[seg] - y[seg - 1]) / pow(t[seg] - t[seg-1], 3)
        if seg < n-1:
            flag = np.array(list(map(lambda x: 1 if x>=t[seg-1] and x<t[seg] else 0, xx)))
        else:
            flag = np.array(list(map(lambda x: 1 if x>=t[seg-1] and x<=t[seg] else 0, xx)))
        yy = yy + (a + b * (xx-t[seg-1]) + c * pow(xx - t[seg-1], 2) + d * pow(xx - t[seg-1], 3)) * flag

    return yy


# Compute today's adjusted implied issuer rating based on the iterative rules
# The adjusted implied issuer rating is calculated according to the defined update logic
def calc_rating_num(df, n):
    alpha = 2/(1 + n)
    if df['rating_num_old'] == 0:
        pass
    else:
        df['rating_num'] = df['rating_num_old'] + alpha * (df['rating_num'] - df['rating_num_old'])
    return df['rating_num']


# Write the result to a new sheet in the Excel file
def excel_add_sheet(dataframe, file_path, sheet_name):
    excelWriter = pd.ExcelWriter(file_path, engine='openpyxl')
    book = load_workbook(excelWriter.path)
    if sheet_name in book.sheetnames:  # If the sheet already exists, delete the old sheet before writing the new one
        idx = book.sheetnames.index(sheet_name)
        book.remove(book.worksheets[idx])
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=False)
    excelWriter.save()
    excelWriter.close()


def df_dif(df1, df2, keys):
    df1 = df1.append(df2)
    df1 = df1.append(df2)
    result = df1.drop_duplicates(subset=keys, keep=False)
    return result


# Preprocessing method for curve data: make_curve
def make_curve(data, rating_num, key_time, key_rate_spread, shift0, bounds):
    sub_data = data[(data['rating_num_final'] == rating_num) & (~data['yi'].isnull())]
    sub_data['exch_flag'] = sub_data['s_info_exchmarket'].apply(lambda r: 1 if '银行间' in r else 0)
    sub_data.sort_values(by=['b_dq_volume', 'exch_flag', 'b_anal_ytm'], ascending=[0, 0, 1], inplace=True)
    sub_data.reset_index(drop=True, inplace=True)
    sub_data['rankvol'] = sub_data.index.values + 1  # 交易点权重排名
    sub_data['weight'] = (max(sub_data['rankvol']) + 1 - sub_data['rankvol']) / (sum(sub_data['rankvol']))  # 交易点权重
    if rating_num < 6:
        xopt = optimize.minimize(fun=myfunc2, x0=shift0, args=[sub_data, key_time, key_rate_spread], bounds=bounds)  # 约束优化求解
        x = xopt.x
        return key_rate_spread + x
    else:
        xopt = optimize.minimize(fun=myfunc3, x0=shift0, args=[sub_data, key_time, key_rate_spread], bounds=bounds)  # 约束优化求解
        x = xopt.x
        return key_rate_spread + x


# Use an optimization function series for the yield-to-maturity (YTM) curve
def myfunc(x, args):
    T, key_time = args
    x = np.cumsum(x)
    calc_ytm = HermiteInter(key_time, x, T['ttm'].values) * 100 + T['yi'].values
    T['calc_ytm'] = None
    T['calc_ytm'] = pd.DataFrame(data=calc_ytm)
    return sum((T['calc_ytm'] - T['b_anal_ytm'])**2 * T['weight'])


def myfunc2(x, args):
    T, key_time, key_rate = args
    x = np.cumsum(key_rate + x)
    calc_ytm = HermiteInter(key_time, x, T['ttm'].values) * 100 + T['yi'].values
    T['calc_ytm'] = None
    T['calc_ytm'] = pd.DataFrame(data=calc_ytm)
    return sum((T['calc_ytm'] - T['b_anal_ytm']) ** 2 * T['weight'])


def myfunc3(x, args):
    T, key_time, key_rate = args
    x = key_rate + x
    calc_ytm = HermiteInter(key_time, x, T['ttm'].values) * 100 + T['yi'].values
    T['calc_ytm'] = None
    T['calc_ytm'] = pd.DataFrame(data=calc_ytm)
    return sum((T['calc_ytm'] - T['b_anal_ytm']) ** 2 * T['weight'])
